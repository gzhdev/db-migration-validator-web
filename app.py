import io
import os

from flask import Flask, render_template, request, redirect, url_for, jsonify, abort, send_file

import db
import importer

app = Flask(__name__)


@app.before_request
def ensure_db():
    db.init_db()


# ---------------------------------------------------------------------------
# Pages
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    runs = db.list_runs()
    return render_template("index.html", runs=runs)


@app.route("/import", methods=["POST"])
def import_run():
    debug_dir = request.form.get("debug_dir", "").strip()
    if not debug_dir:
        return redirect(url_for("index"))
    try:
        run_id = importer.import_debug_dir(debug_dir)
    except ValueError as e:
        runs = db.list_runs()
        return render_template("index.html", runs=runs, error=str(e))
    return redirect(url_for("run_detail", run_id=run_id))


@app.route("/runs/<int:run_id>")
def run_detail(run_id):
    run = db.get_run(run_id)
    if not run:
        abort(404)
    summaries = db.get_table_summaries(run_id)
    return render_template("run_detail.html", run=run, summaries=summaries)


@app.route("/runs/<int:run_id>/records")
def records_page(run_id):
    run = db.get_run(run_id)
    if not run:
        abort(404)
    table_names = db.get_table_names(run_id)
    selected_table = request.args.get("table", "")
    selected_type = request.args.get("match_type", "")
    return render_template(
        "records.html",
        run=run,
        table_names=table_names,
        selected_table=selected_table,
        selected_type=selected_type,
    )


@app.route("/mapping-generator")
def mapping_generator():
    return render_template("mapping_generator.html")


@app.route("/api/excel-template")
def excel_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # -- 样式 --
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    group_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    HEADERS = [
        "source_table", "target_table", "source_schema", "target_schema",
        "description", "source_filter", "target_filter",
        "sample_size", "batch_size",
        "join_tables", "table_filters",
        "target_field", "source_field", "is_primary_key",
        "transform_type", "transform_fields", "transform_params",
        "compare_rule", "tolerance", "nullable",
        "min_value", "max_value", "allowed_values", "pattern",
        "value_check_expr",
    ]

    def style_header(ws):
        for col_idx, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    def add_row(ws, row_num, values, is_group_start=False):
        for col_idx, v in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=v)
            cell.border = thin_border
            if is_group_start:
                cell.fill = group_fill

    def auto_width(ws):
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # ========== Sheet 1: 单表示例 ==========
    ws1 = wb.active
    ws1.title = "单表示例"
    style_header(ws1)

    # 表映射1: users -> user_profile
    rows_s1 = [
        # source_table, target_table, source_schema, target_schema, description,
        # source_filter, target_filter, sample_size, batch_size,
        # join_tables, table_filters,
        # target_field, source_field, is_primary_key,
        # transform_type, transform_fields, transform_params,
        # compare_rule, tolerance, nullable,
        # min_value, max_value, allowed_values, pattern, value_check_expr
        ["users", "user_profile", "", "", "用户表迁移",
         "deleted_at IS NULL", "is_deleted = false", 10000, 500,
         "", "",
         "user_id", "id", "true",
         "", "", "",
         "", "", "",
         "", "", "", "", ""],
        ["", "", "", "", "",
         "", "", "", "",
         "", "",
         "full_name", "", "",
         "concat", "first_name|last_name", '{"separator": " "}',
         "", "", "",
         "", "", "", "", ""],
        ["", "", "", "", "",
         "", "", "", "",
         "", "",
         "email_address", "email", "",
         "", "", "",
         "ignore_case", "", "",
         "", "", "", "", ""],
        ["", "", "", "", "",
         "", "", "", "",
         "", "",
         "balance", "account_balance", "",
         "", "", "",
         "numeric_tolerance", 0.01, "",
         "", "", "", "", ""],
    ]
    for i, row in enumerate(rows_s1):
        add_row(ws1, i + 2, row, is_group_start=(i == 0))
    auto_width(ws1)

    # ========== Sheet 2: 多表 JOIN 示例 ==========
    ws2 = wb.create_sheet("多表JOIN示例")
    style_header(ws2)

    join_tables_json = '[{"table_name":"orders","alias":"o","join_type":"primary"},{"table_name":"users","alias":"u","join_type":"left","join_condition":"o.user_id = u.id"}]'
    table_filters_json = '{"o":"o.deleted_at IS NULL"}'

    rows_s2 = [
        ["", "order_detail", "", "", "订单明细（多表JOIN）",
         "", "", 0, 1000,
         join_tables_json, table_filters_json,
         "order_id", "o.id", "true",
         "", "", "",
         "", "", "",
         "", "", "", "", ""],
        ["", "", "", "", "",
         "", "", "", "",
         "", "",
         "customer", "u.name", "",
         "", "", "",
         "", "", "true",
         "", "", "", "", ""],
        ["", "", "", "", "",
         "", "", "", "",
         "", "",
         "total", "", "",
         "math", "o.price|o.qty", '{"expression": "o.price * o.qty"}',
         "numeric_tolerance", 0.01, "",
         "", "", "", "", ""],
    ]
    for i, row in enumerate(rows_s2):
        add_row(ws2, i + 2, row, is_group_start=(i == 0))
    auto_width(ws2)

    # ========== Sheet 3: 多源插入同一目标表示例 ==========
    ws3 = wb.create_sheet("多源插入示例")
    style_header(ws3)

    join_json_s3 = '[{"table_name":"table_a","alias":"a","join_type":"primary"},{"table_name":"table_b","alias":"b","join_type":"left","join_condition":"a.id = b.ref_id"}]'

    rows_s3 = [
        # 条目1: 单表 table_a -> target (条件1)
        ["table_a", "target", "", "", "条件1: 来自A表 (type=X)",
         "type = 'X'", "type = 'X'", 0, 1000,
         "", "",
         "id", "id", "true",
         "", "", "",
         "", "", "",
         "", "", "", "", ""],
        ["", "", "", "", "",
         "", "", "", "",
         "", "",
         "name", "name", "",
         "", "", "",
         "", "", "",
         "", "", "", "", ""],
        ["", "", "", "", "",
         "", "", "", "",
         "", "",
         "value", "amount", "",
         "cast", "amount", '{"to": "int"}',
         "", "", "",
         "", "", "", "", ""],
        # 条目2: 多表 JOIN -> target (条件2)
        ["", "target", "", "", "条件2: 来自A JOIN B (type!=X, category=Y)",
         "", "type != 'X' AND category = 'Y'", 0, 1000,
         join_json_s3, '{"a":"a.type != \'X\' AND a.category = \'Y\'"}',
         "id", "a.id", "true",
         "", "", "",
         "", "", "",
         "", "", "", "", ""],
        ["", "", "", "", "",
         "", "", "", "",
         "", "",
         "name", "b.display_name", "",
         "", "", "",
         "", "", "true",
         "", "", "", "", ""],
        ["", "", "", "", "",
         "", "", "", "",
         "", "",
         "value", "", "",
         "math", "a.price|b.factor", '{"expression": "a.price * b.factor"}',
         "numeric_tolerance", 0.01, "",
         "", "", "", "", ""],
        # 条目3: 单表 table_c -> target (条件3)
        ["table_c", "target", "", "", "条件3: 来自C表 (type!=X, category!=Y)",
         "type != 'X' AND category != 'Y'", "type != 'X' AND category != 'Y'", 0, 1000,
         "", "",
         "id", "id", "true",
         "", "", "",
         "", "", "",
         "", "", "", "", ""],
        ["", "", "", "", "",
         "", "", "", "",
         "", "",
         "name", "title", "",
         "upper", "title", "",
         "", "", "",
         "", "", "", "", ""],
        ["", "", "", "", "",
         "", "", "", "",
         "", "",
         "value", "score", "",
         "", "", "",
         "numeric_tolerance", 1, "",
         "", "", "", "", ""],
    ]
    for i, row in enumerate(rows_s3):
        is_group = row[0] or row[1]  # new group when source_table or target_table present
        add_row(ws3, i + 2, row, is_group_start=bool(is_group))
    auto_width(ws3)

    # ========== Sheet 4: 填写说明 ==========
    ws4 = wb.create_sheet("填写说明")
    instructions = [
        ["列名", "说明", "适用模式"],
        ["source_table", "源表名（单表模式填写）", "单表"],
        ["target_table", "目标表名（每个映射条目的第一行填写）", "通用"],
        ["source_schema", "源表 Schema（可选）", "单表"],
        ["target_schema", "目标表 Schema（可选）", "通用"],
        ["description", "映射说明（可选，在条目第一行填写）", "通用"],
        ["source_filter", "源表过滤条件（单表模式）", "单表"],
        ["target_filter", "目标表过滤条件（多源插入同一目标表时必填，条件须互斥且完整覆盖）", "通用"],
        ["sample_size", "抽样数量（0=全量）", "通用"],
        ["batch_size", "批量大小（默认1000）", "通用"],
        ["join_tables", "多表JOIN配置（JSON数组），填写后自动切换为多表模式", "多表JOIN"],
        ["table_filters", "各源表过滤条件（JSON对象，key为别名）", "多表JOIN"],
        ["target_field", "目标表字段名（每行一个字段）", "通用"],
        ["source_field", "源表字段名（无转换时填写；多表用 alias.field 格式）", "通用"],
        ["is_primary_key", "是否主键（true/false）", "通用"],
        ["transform_type", "转换类型: concat, upper, lower, trim, substring, replace, constant, coalesce, case, date_format, json_extract, cast, math", "通用"],
        ["transform_fields", "转换源字段（多个用 | 分隔）", "通用"],
        ["transform_params", "转换参数（JSON格式）", "通用"],
        ["compare_rule", "比较规则: exact, ignore_case, ignore_whitespace, numeric_tolerance, skip", "通用"],
        ["tolerance", "数值容差（compare_rule=numeric_tolerance时填写）", "通用"],
        ["nullable", "是否可空（默认true；填false表示NULL不通过校验）", "通用"],
        ["min_value", "最小值校验", "通用"],
        ["max_value", "最大值校验", "通用"],
        ["allowed_values", "允许值列表（用 | 分隔）", "通用"],
        ["pattern", "正则表达式校验", "通用"],
        ["value_check_expr", "自定义Spark SQL校验表达式", "通用"],
        [],
        ["【多源插入同一目标表】使用说明："],
        ["当存储过程有多个INSERT语句从不同源表插入同一目标表时，创建多个映射条目（每个INSERT对应一个条目），"],
        ["每个条目指向相同的target_table，但配置不同的source_table/join_tables和互斥的target_filter。"],
        ["详见「多源插入示例」Sheet。"],
    ]
    for r_idx, row in enumerate(instructions, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
    for col in ws4.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws4.column_dimensions[col[0].column_letter].width = min(max_len + 4, 80)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="mapping_template.xlsx",
    )


# ---------------------------------------------------------------------------
# JSON API
# ---------------------------------------------------------------------------

@app.route("/api/runs/<int:run_id>/records")
def api_records(run_id):
    run = db.get_run(run_id)
    if not run:
        abort(404)
    table = request.args.get("table", "")
    match_type = request.args.get("match_type", "")
    pk_search = request.args.get("pk_search", "").strip()
    try:
        page = int(request.args.get("page", 1))
        page_size = int(request.args.get("page_size", 20))
    except ValueError:
        page, page_size = 1, 20

    result = db.query_records(
        run_id,
        table_name=table or None,
        match_type=match_type or None,
        pk_search=pk_search or None,
        page=page,
        page_size=page_size,
    )
    return jsonify(result)


@app.route("/runs/<int:run_id>", methods=["DELETE"])
def delete_run(run_id):
    run = db.get_run(run_id)
    if not run:
        abort(404)
    db.delete_run(run_id)
    return jsonify({"ok": True})


if __name__ == "__main__":
    db.init_db()
    app.run(debug=True, port=5000)
